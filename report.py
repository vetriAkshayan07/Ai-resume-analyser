import os
import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.units import mm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import Config
from models.database import get_connection


def _ensure_reports_dir():
    os.makedirs(Config.REPORTS_FOLDER, exist_ok=True)


def _get_candidates_for_report(jd_id: int = None) -> list:
    """Fetch all candidates (optionally filtered by JD) from the database."""
    conn = get_connection()
    try:
        cursor = conn.cursor()
        if jd_id:
            cursor.execute(
                "SELECT * FROM candidates WHERE job_description_id = ? ORDER BY rank ASC",
                (jd_id,)
            )
        else:
            cursor.execute("SELECT * FROM candidates ORDER BY rank ASC")
        return [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# PDF Report
# ---------------------------------------------------------------------------

def generate_pdf_report(admin_id: int, jd_id: int = None) -> str:
    """Generate a PDF report and return the file path."""
    _ensure_reports_dir()
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"resume_report_{timestamp}.pdf"
    filepath = os.path.join(Config.REPORTS_FOLDER, filename)

    candidates = _get_candidates_for_report(jd_id)
    doc = SimpleDocTemplate(filepath, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=20,
        textColor=colors.HexColor("#1a1a2e"),
        spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=colors.HexColor("#16213e"),
        spaceAfter=4,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
    )
    summary_style = ParagraphStyle(
        "Summary",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#444444"),
    )

    elements = []

    # Title
    elements.append(Paragraph("AI Resume Screening – Candidate Report", title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.datetime.now().strftime('%B %d, %Y  %H:%M')}  |  Total Candidates: {len(candidates)}",
        body_style
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#4361ee")))
    elements.append(Spacer(1, 6*mm))

    if not candidates:
        elements.append(Paragraph("No candidates found.", body_style))
    else:
        for c in candidates:
            # Candidate heading
            elements.append(Paragraph(
                f"#{c.get('rank', '–')}  {c.get('name', 'Unknown')}  —  Match: {c.get('match_score', 0):.1f}%",
                heading_style
            ))

            # Info table
            data = [
                ["Email", c.get("email", "N/A"), "Phone", c.get("phone", "N/A")],
                ["Location", c.get("location", "N/A"), "Experience", c.get("experience", "N/A")],
                ["Education", Paragraph(c.get("education", "N/A")[:120], body_style), "Certifications", Paragraph(c.get("certifications", "N/A")[:120], body_style)],
                ["Skills", Paragraph(c.get("skills", "N/A")[:200], body_style), "", ""],
            ]
            table = Table(data, colWidths=[35*mm, 65*mm, 35*mm, 55*mm])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8ecff")),
                ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#e8ecff")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
                ("PADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)

            # AI Summary
            summary = c.get("ai_summary") or "No AI summary available."
            elements.append(Spacer(1, 2*mm))
            elements.append(Paragraph("<b>AI Summary:</b> " + summary[:400], summary_style))
            elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dddddd")))
            elements.append(Spacer(1, 4*mm))

    doc.build(elements)

    # Save report record
    conn = get_connection()
    conn.execute(
        "INSERT INTO reports (report_type, filepath, generated_by) VALUES (?, ?, ?)",
        ("pdf", filepath, admin_id)
    )
    conn.commit()
    conn.close()

    return filepath


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

def generate_excel_report(admin_id: int, jd_id: int = None) -> str:
    """Generate an Excel report and return the file path."""
    _ensure_reports_dir()
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"resume_report_{timestamp}.xlsx"
    filepath = os.path.join(Config.REPORTS_FOLDER, filename)

    candidates = _get_candidates_for_report(jd_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Rankings"

    # Colors
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    alt_fill = PatternFill(start_color="EEF0FF", end_color="EEF0FF", fill_type="solid")
    border_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Header row
    headers = [
        "Rank", "Name", "Email", "Phone", "Location",
        "Skills", "Experience", "Education", "Certifications",
        "Match Score (%)", "AI Summary"
    ]
    col_widths = [6, 20, 25, 15, 15, 35, 15, 30, 25, 14, 60]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, c in enumerate(candidates, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        row_data = [
            c.get("rank", ""),
            c.get("name", ""),
            c.get("email", ""),
            c.get("phone", ""),
            c.get("location", ""),
            c.get("skills", ""),
            c.get("experience", ""),
            c.get("education", ""),
            c.get("certifications", ""),
            c.get("match_score", 0),
            c.get("ai_summary", ""),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            if fill.fill_type:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 60

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(filepath)

    # Save report record
    conn = get_connection()
    conn.execute(
        "INSERT INTO reports (report_type, filepath, generated_by) VALUES (?, ?, ?)",
        ("excel", filepath, admin_id)
    )
    conn.commit()
    conn.close()

    return filepath
